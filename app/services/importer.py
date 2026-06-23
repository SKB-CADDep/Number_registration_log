"""
Модуль сервиса импорта исторических данных из Excel.

Отвечает за парсинг и загрузку legacy-данных (старых регистрационных журналов)
в систему. Создаёт записи документов, оборудования, пользователей и помечает
номера как уже назначенные.

Важные особенности текущей реализации:
- Каждое оборудование создаётся заново по каждой строке (дублирование возможно).
- Пользователи создаются/переиспользуются по полю "Имя пользователя в системе".
- После импорта сдвигается глобальный счётчик номеров (`counter`), чтобы новые
  регистрации не пересекались с импортированными.
- Импорт работает в рамках одной транзакции — при ошибке на любой строке
  откатывается всё.
"""

from __future__ import annotations

from datetime import datetime

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from openpyxl import load_workbook

from app.repositories.documents import DocumentsRepository
from app.repositories.equipment import EquipmentRepository
from app.repositories.users import UsersRepository
from app.repositories.doc_numbers import DocNumbersRepository
from app.repositories.counter import CounterRepository
from app.models.doc_number import DocNumStatus


class ExcelImporterService:
    """
    Сервис для пакетного импорта данных из Excel-файлов.

    Предназначен для первоначальной миграции данных из унаследованных систем.
    Обрабатывает файл построчно, создавая необходимые сущности и поддерживая
    целостность нумерации.
    """

    def __init__(self, session: AsyncSession) -> None:
        """Инициализация сервиса со всеми необходимыми репозиториями."""
        self.session = session
        self.docs_repo = DocumentsRepository(session)
        self.eq_repo = EquipmentRepository(session)
        self.users_repo = UsersRepository(session)
        self.docnums_repo = DocNumbersRepository(session)
        self.counter_repo = CounterRepository(session)

    async def import_file(self, path: str) -> dict:
        """
        Выполняет импорт данных из Excel-файла.

        Ожидаемая структура колонок (заголовки первой строки):
        - № документа
        - Дата регистрации
        - Наименование документа
        - Примечание
        - Тип оборудования
        - № заводской
        - № заказа
        - Маркировка
        - № станционный
        - Станция / Объект
        - Фамилия
        - Имя
        - Отчество
        - Отдел
        - Имя пользователя в системе

        Логика обработки одной строки:
            1. Парсит номер документа → извлекает числовую часть.
            2. Создаёт/находит пользователя по логину.
            3. Создаёт новую запись оборудования (даже если такое уже есть).
            4. Создаёт документ.
            5. Добавляет запись в `doc_numbers` со статусом `assigned`.
            6. Отслеживает максимальный номер для сдвига счётчика.

        После обработки всех строк обновляет глобальный счётчик номеров.

        Args:
            path: Полный путь к Excel-файлу (.xlsx).

        Returns:
            dict: Статистика импорта:
                - `imported`: количество успешно импортированных строк
                - `next_start`: следующее значение счётчика (max_numeric + 1)

        Raises:
            Различные исключения при ошибках чтения файла или работы с БД.
            Рекомендуется обораживать вызов в try-except на уровне вызывающего кода.
        """
        wb = load_workbook(filename=path)
        ws = wb.active

        # Ожидаемые заголовки (порядок не важен, ищется по названию)
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers)}

        max_numeric = 0
        created = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            doc_no = row[col.get("№ документа")]
            if not doc_no:
                continue

            numeric = int(str(doc_no).split("-")[-1])
            reg_date = row[col.get("Дата регистрации")]
            if isinstance(reg_date, str):
                try:
                    reg_date = datetime.strptime(reg_date, "%d.%m.%Y")
                except Exception:
                    try:
                        reg_date = datetime.fromisoformat(reg_date)
                    except Exception:
                        reg_date = datetime.now()

            doc_name = row[col.get("Наименование документа")]
            note = row[col.get("Примечание")]
            eq_type = row[col.get("Тип оборудования")]
            factory_no = row[col.get("№ заводской")]
            order_no = row[col.get("№ заказа")]
            label = row[col.get("Маркировка")]
            station_no = row[col.get("№ станционный")]
            station_object = row[col.get("Станция / Объект")]

            last_name = row[col.get("Фамилия")] if "Фамилия" in col else None
            first_name = row[col.get("Имя")] if "Имя" in col else None
            middle_name = row[col.get("Отчество")] if "Отчество" in col else None
            department = row[col.get("Отдел")] if "Отдел" in col else None
            username = row[col.get("Имя пользователя в системе")] if "Имя пользователя в системе" in col else None
            username = username or "import"

            # Создание/получение пользователя
            user = await self.users_repo.get_by_username(username)
            if not user:
                user = await self.users_repo.create(username)
                user.last_name = last_name
                user.first_name = first_name
                user.middle_name = middle_name
                user.department = department

            # Создание оборудования (упрощённая логика — создаётся каждый раз)
            eq = await self.eq_repo.create(
                {
                    "eq_type": eq_type or "N/A",
                    "factory_no": factory_no,
                    "order_no": order_no,
                    "label": label,
                    "station_no": station_no,
                    "station_object": station_object,
                    "notes": None,
                }
            )

            # Создание документа
            doc = await self.docs_repo.create(
                {
                    "numeric": numeric,
                    "reg_date": reg_date,
                    "doc_name": doc_name or "",
                    "note": note or "",
                    "equipment_id": eq.id,
                    "user_id": user.id,
                }
            )

            # Помечаем номер как назначенный
            from app.models.doc_number import DocNumber

            dn = DocNumber(
                numeric=numeric,
                is_golden=(numeric % 100 == 0),
                status=DocNumStatus.assigned,
                reserved_by=None,
                session_id=None,
                assigned_at=reg_date,
            )
            self.session.add(dn)

            max_numeric = max(max_numeric, numeric)
            created += 1

        # Обновляем глобальный счётчик номеров
        if max_numeric > 0:
            await self.counter_repo.set_after_import(max_numeric + 1)

        await self.session.commit()

        return {
            "imported": created,
            "next_start": max_numeric + 1 if max_numeric > 0 else 1,
        }

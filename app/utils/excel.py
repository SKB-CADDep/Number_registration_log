"""
Модуль генерации Excel-отчётов.

Содержит класс `ReportExcelBuilder`, который отвечает за создание
файлов Excel (.xlsx) на основе данных отчётов.

Поддерживает три типа отчётов:
- Базовый отчёт (устаревающий)
- Расширенный отчёт (основной, используемый фронтендом)
- Административный отчёт (с техническими идентификаторами для редактирования)

Все отчёты сохраняются в директорию `var/exports/`.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os


class ReportExcelBuilder:
    """
    Построитель Excel-отчётов на основе данных из `ReportsService`.

    Предоставляет три метода генерации отчётов разной детализации.
    Использует библиотеку `openpyxl` для программного создания XLSX-файлов.
    """

    # Заголовки для базового отчёта (с ФИО по отдельности)
    columns = [
        "№ документа",
        "Дата регистрации",
        "Наименование документа",
        "Примечание",
        "Тип оборудования",
        "№ заводской",
        "№ заказа",
        "Маркировка",
        "№ станционный",
        "Станция / Объект",
        "Фамилия",
        "Имя",
        "Отчество",
        "Отдел",
    ]

    # Заголовки для расширенного отчёта
    columns_extended = [
        "№ документа",
        "Дата регистрации",
        "Наименование документа",
        "Примечание",
        "Тип оборудования",
        "№ заводской",
        "№ заказа",
        "Маркировка",
        "№ станционный",
        "Станция / Объект",
        "Пользователь (создавший)",
    ]

    def build_report(self, rows: list[dict]) -> str:
        """
        Создаёт базовый Excel-отчёт (устаревший формат).

        Используется для совместимости со старым кодом. Формирует отчёт
        с разбивкой пользователя на ФИО + отдел.

        Args:
            rows: Список словарей из `ReportsService.get_rows()`.

        Returns:
            str: Путь к созданному Excel-файлу.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"

        ws.append(self.columns)

        for r in rows:
            ws.append(
                [
                    r["doc_no"],
                    r["reg_date"],
                    r["doc_name"],
                    r["note"],
                    r["eq_type"],
                    r["factory_no"],
                    r["order_no"],
                    r["label"],
                    r["station_no"],
                    r["station_object"],
                    r["last_name"] or (r.get("username_fallback") or ""),
                    r.get("first_name") or "",
                    r.get("middle_name") or "",
                    r.get("department") or "",
                ]
            )

        # Авто-ширина колонок
        for i, col_name in enumerate(self.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(len(col_name) + 2, 15)

        Path("var/exports").mkdir(parents=True, exist_ok=True)
        fname = f"var/exports/report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(fname)
        return fname

    def build_report_extended(self, rows: list[dict]) -> str:
        """
        Создаёт расширенный Excel-отчёт (основной формат).

        Используется для большинства пользовательских запросов на экспорт.
        Содержит упрощённую колонку "Пользователь (создавший)" вместо ФИО.

        Args:
            rows: Данные из `ReportsService.get_rows_extended()`.

        Returns:
            str: Путь к созданному Excel-файлу.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"

        # Заголовки
        for col, header in enumerate(self.columns_extended, 1):
            ws.cell(row=1, column=col, value=header)

        # Данные
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(self.columns_extended, 1):
                value = {
                    "№ документа": row_data["doc_no"],
                    "Дата регистрации": row_data["reg_date"],
                    "Наименование документа": row_data["doc_name"],
                    "Примечание": row_data["note"],
                    "Тип оборудования": row_data["eq_type"],
                    "№ заводской": row_data["factory_no"],
                    "№ заказа": row_data["order_no"],
                    "Маркировка": row_data["label"],
                    "№ станционный": row_data["station_no"],
                    "Станция / Объект": row_data["station_object"],
                    "Пользователь (создавший)": row_data["username"],
                }.get(col_name)

                ws.cell(row=row_idx, column=col_idx, value=value)

        self._auto_adjust_column_width(ws)

        filename = f"var/exports/report_extended_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb.save(filename)
        return filename

    def build_report_extended_admin(self, rows: list[dict]) -> str:
        """
        Создаёт административный Excel-отчёт с расширенными колонками.

        Включает технические поля (`ID документа`, `numeric` и т.д.),
        необходимые для административной панели и перехода к редактированию.

        Args:
            rows: Данные из `ReportsService.get_rows_extended_admin()`.

        Returns:
            str: Путь к созданному Excel-файлу.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Админский отчет"

        admin_columns = [
            "ID документа",
            "№ документа",
            "Дата регистрации",
            "Наименование документа",
            "Примечание",
            "Тип оборудования",
            "№ заводской",
            "№ заказа",
            "Маркировка",
            "№ станционный",
            "Станция / Объект",
            "Пользователь (создавший)",
        ]

        # Заголовки
        for col, header in enumerate(admin_columns, 1):
            ws.cell(row=1, column=col, value=header)

        # Данные
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(admin_columns, 1):
                value = {
                    "ID документа": row_data.get("id"),
                    "№ документа": row_data.get("doc_no"),
                    "Дата регистрации": row_data.get("reg_date"),
                    "Наименование документа": row_data.get("doc_name"),
                    "Примечание": row_data.get("note"),
                    "Тип оборудования": row_data.get("eq_type"),
                    "№ заводской": row_data.get("factory_no"),
                    "№ заказа": row_data.get("order_no"),
                    "Маркировка": row_data.get("label"),
                    "№ станционный": row_data.get("station_no"),
                    "Станция / Объект": row_data.get("station_object"),
                    "Пользователь (создавший)": row_data.get("username"),
                }.get(col_name)

                ws.cell(row=row_idx, column=col_idx, value=value)

        self._auto_adjust_column_width(ws)

        filename = f"var/exports/report_admin_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb.save(filename)
        return filename

    def _auto_adjust_column_width(self, ws: Workbook.worksheets[0]) -> None:
        """
        Автоматически подбирает ширину колонок по содержимому.

        Максимальная ширина ограничена 50 символами.
        Используется в расширенных и административных отчётах.
        """
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    continue
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
